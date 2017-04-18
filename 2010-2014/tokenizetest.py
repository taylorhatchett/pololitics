import openpyxl

wb = openpyxl.load_workbook("/Users/thatchett/Documents/Jokes Analysis/2010-2014/forNLTKAnalysis.xlsx")
ws = wb.get_sheet_by_name('Sheet2')

import nltk
from nltk import word_tokenize, pos_tag
from nltk.corpus import stopwords

stop_words = set(stopwords.words("english"))
print(stop_words)
'''filter = []

for i in range(2, 12065):
    count = 0
    f_joke = ws.cell(row = i, column = 4)
    joke = f_joke.value

    import re
    letters_only = re.sub("[^a-zA-Z0-9]", " ", joke)
    joke1 = letters_only
    joke1 = ' '.join([word for word in joke1.split() if word not in stop_words])
#   print(joke1)
    ws.cell(row = i, column = 5).value = joke1
	
wb.save('/Users/thatchett/Documents/Jokes Analysis/2010-2014/forNLTKAnalysis1.xlsx')
'''
